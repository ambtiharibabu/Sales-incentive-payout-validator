import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Load validated data ---
df = pd.read_csv("data/processed/validated_participants.csv")

wb = Workbook()

# -------------------------------------------------------
# SHEET 1: All Participants with validation flags
# -------------------------------------------------------
ws1 = wb.active
ws1.title = "All Participants"

# Define red fill for flagged rows
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
header_font = Font(bold=True)

# Write header row
headers = list(df.columns)
ws1.append(headers)

# Bold the header
for cell in ws1[1]:
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Write data rows — apply red fill if overall_status == "Review"
for row in dataframe_to_rows(df, index=False, header=False):
    ws1.append(row)

# Color rows after header (row 1)
for row_idx, status in enumerate(df["overall_status"], start=2):  # start=2 skips header
    if status == "Review":
        for cell in ws1[row_idx]:
            cell.fill = red_fill

# -------------------------------------------------------
# SHEET 2: Discrepancies Only
# -------------------------------------------------------
ws2 = wb.create_sheet("Discrepancies Only")

df_disc = df[df["discrepancy_flag"] == True].copy()  # filter discrepancy rows only

disc_headers = list(df_disc.columns)
ws2.append(disc_headers)

for cell in ws2[1]:
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for row in dataframe_to_rows(df_disc, index=False, header=False):
    ws2.append(row)

# -------------------------------------------------------
# SHEET 3: LOB Summary pivot
# -------------------------------------------------------
ws3 = wb.create_sheet("LOB Summary")

# Build summary: count, total approved payout, avg attainment per LOB
summary = df.groupby("lob").agg(
    participant_count=("participant_id", "count"),
    total_payout_approved=("payout_approved_usd", "sum"),
    avg_attainment_pct=("attainment_pct", "mean"),
    discrepancy_count=("discrepancy_flag", "sum"),
    review_count=("overall_status", lambda x: (x == "Review").sum())
).reset_index()

summary["avg_attainment_pct"] = summary["avg_attainment_pct"].round(4)
summary["total_payout_approved"] = summary["total_payout_approved"].round(2)

summary_headers = list(summary.columns)
ws3.append(summary_headers)

for cell in ws3[1]:
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for row in dataframe_to_rows(summary, index=False, header=False):
    ws3.append(row)

# -------------------------------------------------------
# Save workbook
# -------------------------------------------------------
output_path = "outputs/validated_payouts.xlsx"
wb.save(output_path)

print(f"✅ Excel saved → {output_path}")
print(f"   Sheet 1: {len(df)} rows (all participants)")
print(f"   Sheet 2: {len(df_disc)} rows (discrepancies only)")
print(f"   Sheet 3: {len(summary)} LOBs summarized")